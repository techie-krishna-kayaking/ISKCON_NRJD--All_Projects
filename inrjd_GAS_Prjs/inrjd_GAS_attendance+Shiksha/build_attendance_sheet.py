"""
Reads tab3 (raw attendance events) from attendance-db.xlsx and creates a new
'attendance' sheet with aggregated per-(PROGRAM_KEY, DEVOTEE) summaries matching
the unified project's schema:

  PROGRAM KEY | AREA | SUB_AREA | FREQUENCY | TYPE_OF_PROGRAM | LANGUAGE |
  PROGRAM_OWNER | DEVOTEE | TOTAL_SESSIONS | ATTENDED | PERCENTAGE |
  LAST_ATT_DATE | HOST_NAME | BV_CHAPTER | LAST_UPDATED

Aggregation logic (matches the old updateTab4Summary):
  - TOTAL_SESSIONS = number of unique dates for that program key
  - ATTENDED = number of rows where this devotee was 'present' for that program
  - PERCENTAGE = round(ATTENDED / TOTAL_SESSIONS * 100)%
  - LAST_ATT_DATE = most recent DATE for that (program, devotee) pair
  - HOST_NAME = host from the most recent session
  - BV_CHAPTER = chapter from the most recent session
  - LAST_UPDATED = timestamp of this script run
"""

import openpyxl
from collections import defaultdict
from datetime import datetime

XLSX_PATH = 'attendance-db.xlsx'

# ── Column indices in tab3 ──
T3_PROGRAM_KEY = 0
T3_AREA = 1
T3_SUB_AREA = 2
T3_FREQUENCY = 3
T3_TYPE = 4
T3_LANGUAGE = 5
T3_OWNER = 6
T3_CHAPTER = 13   # BV_CHAPTER
T3_HOST = 14      # HOST
T3_DEVOTEE = 15
T3_DATE = 16
T3_STATUS = 17    # 'present' or 'absent'

# New attendance sheet headers
ATT_HEADERS = [
    'PROGRAM KEY', 'AREA', 'SUB_AREA', 'FREQUENCY', 'TYPE_OF_PROGRAM',
    'LANGUAGE', 'PROGRAM_OWNER', 'DEVOTEE', 'TOTAL_SESSIONS', 'ATTENDED',
    'PERCENTAGE', 'LAST_ATT_DATE', 'HOST_NAME', 'BV_CHAPTER', 'LAST_UPDATED'
]


def parse_date(val):
    """Try to parse a date value into a comparable date object."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def format_date(dt):
    """Format a datetime to YYYY-MM-DD string."""
    if dt is None:
        return ''
    return dt.strftime('%Y-%m-%d')


def main():
    print(f'Opening {XLSX_PATH}...')
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    # ── Read tab3 ──
    ws3 = wb['tab3']
    tab3_rows = list(ws3.iter_rows(values_only=True))
    print(f'tab3: {len(tab3_rows)} rows (including header)')

    if len(tab3_rows) < 2:
        print('No data in tab3!')
        return

    # ── Read tab1 for program metadata (in case tab3 has stale metadata) ──
    ws1 = wb['tab1']
    tab1_rows = list(ws1.iter_rows(values_only=True))
    program_meta = {}  # programKey → {area, sub_area, freq, type, lang, owner}
    for row in tab1_rows[1:]:
        pk = str(row[0] or '').strip()
        if pk:
            program_meta[pk] = {
                'area': str(row[1] or '').strip(),
                'sub_area': str(row[2] or '').strip(),
                'frequency': str(row[3] or '').strip(),
                'type': str(row[4] or '').strip(),
                'language': str(row[5] or '').strip(),
                'owner': str(row[6] or '').strip(),
            }
    print(f'tab1: {len(program_meta)} programs loaded')

    # ── Aggregate tab3 ──
    # Per program: track unique session dates
    program_dates = defaultdict(set)  # programKey → set of dates

    # Per (programKey, devotee): track attendance
    DevKey = tuple  # (programKey, devotee)
    devotee_data = defaultdict(lambda: {
        'present_dates': set(),   # unique dates marked present
        'all_dates': set(),       # unique dates this devotee appeared (present or absent)
        'total_rows': 0,
        'last_date': None,
        'last_host': '',
        'last_chapter': '',
        # Keep program metadata from the row as fallback
        'area': '', 'sub_area': '', 'frequency': '', 'type': '',
        'language': '', 'owner': ''
    })

    skipped = 0
    for i, row in enumerate(tab3_rows[1:], start=2):
        pk = str(row[T3_PROGRAM_KEY] or '').strip()
        devotee = str(row[T3_DEVOTEE] or '').strip()
        status = str(row[T3_STATUS] or '').strip().lower()
        date_val = row[T3_DATE]
        host = str(row[T3_HOST] or '').strip()
        chapter = str(row[T3_CHAPTER] or '').strip()

        if not pk or not devotee:
            skipped += 1
            continue

        dt = parse_date(date_val)
        date_key = format_date(dt) if dt else str(date_val or '')

        # Track unique session dates per program
        if date_key:
            program_dates[pk].add(date_key)

        key = (pk, devotee)
        d = devotee_data[key]
        d['total_rows'] += 1

        # Track unique dates for this devotee
        if date_key:
            d['all_dates'].add(date_key)
            if status == 'present':
                d['present_dates'].add(date_key)

        # Track most recent session for this devotee
        if dt and (d['last_date'] is None or dt > d['last_date']):
            d['last_date'] = dt
            d['last_host'] = host
            d['last_chapter'] = chapter

        # Fallback metadata from tab3 row
        if not d['area']:
            d['area'] = str(row[T3_AREA] or '').strip()
            d['sub_area'] = str(row[T3_SUB_AREA] or '').strip()
            d['frequency'] = str(row[T3_FREQUENCY] or '').strip()
            d['type'] = str(row[T3_TYPE] or '').strip()
            d['language'] = str(row[T3_LANGUAGE] or '').strip()
            d['owner'] = str(row[T3_OWNER] or '').strip()

    print(f'Processed {len(tab3_rows)-1} rows, skipped {skipped} empty rows')
    print(f'Unique (program, devotee) pairs: {len(devotee_data)}')
    print(f'Unique programs with sessions: {len(program_dates)}')

    # ── Build output rows ──
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    output_rows = []

    for (pk, devotee), d in sorted(devotee_data.items()):
        total_sessions = len(program_dates.get(pk, set()))
        attended = len(d['present_dates'])
        percentage = round((attended / total_sessions) * 100) if total_sessions > 0 else 0

        # Use tab1 metadata if available, otherwise fall back to tab3
        meta = program_meta.get(pk, {})
        area = meta.get('area', d['area'])
        sub_area = meta.get('sub_area', d['sub_area'])
        frequency = meta.get('frequency', d['frequency'])
        prog_type = meta.get('type', d['type'])
        language = meta.get('language', d['language'])
        owner = meta.get('owner', d['owner'])

        output_rows.append([
            pk,
            area,
            sub_area,
            frequency,
            prog_type,
            language,
            owner,
            devotee,
            total_sessions,
            attended,
            f'{percentage}%',
            format_date(d['last_date']),
            d['last_host'],
            d['last_chapter'],
            now_str
        ])

    print(f'Generated {len(output_rows)} attendance rows')

    # ── Write to new 'attendance' sheet ──
    # Remove existing attendance sheet if present
    if 'attendance' in wb.sheetnames:
        del wb['attendance']
        print('Removed existing "attendance" sheet')

    ws_att = wb.create_sheet('attendance', 0)  # Insert at position 0

    # Write header
    ws_att.append(ATT_HEADERS)

    # Write data
    for row in output_rows:
        ws_att.append(row)

    # ── Save ──
    # Need to reopen in write mode since read_only won't work
    wb.close()

    # Reopen in full mode and write
    print('Reopening in write mode...')
    wb2 = openpyxl.load_workbook(XLSX_PATH)

    if 'attendance' in wb2.sheetnames:
        del wb2['attendance']

    ws_new = wb2.create_sheet('attendance', 0)
    ws_new.append(ATT_HEADERS)
    for row in output_rows:
        ws_new.append(row)

    wb2.save(XLSX_PATH)
    wb2.close()

    print(f'\nDone! "attendance" sheet created with {len(output_rows)} rows + header.')
    print(f'File saved: {XLSX_PATH}')
    print('\nNext steps:')
    print('  1. Open attendance-db.xlsx')
    print('  2. Go to the "attendance" sheet')
    print('  3. Select all data (Ctrl+A)')
    print('  4. Copy (Ctrl+C)')
    print('  5. Go to your Google Sheet → "attendance" tab')
    print('  6. Clear existing data, paste (Ctrl+V)')


if __name__ == '__main__':
    main()
